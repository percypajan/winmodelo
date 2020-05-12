import openpyxl

from .models import *
from .variables  import *
from datetime import date

def BDgrabaregistro(login,miper,cod,tipo, barra, potencia,data): # tipo=1: Registro  tipo=2: Con.Libre3eros  tipo=3: Reparto

    reg=Registro.objects.filter(mesano=miper, codigo=cod, tipo=tipo,  barra=barra, version=0, potencia=potencia) #borrara todos los registros de todos dias del mes
    
    reg.delete()

    cont=1
    for dia in range(31):
        reg=Registro()
        reg.mesano_id  =miper
        reg.codigo  =cod
        reg.tipo     =tipo
        reg.barra    =barra
        reg.potencia=potencia
        reg.version  = 0
        reg.dia=dia+1                    
        reg.descripcion=str(miper)+'-'+str(tipo)+'-'+str(0) +'-'+data[0]+'-'+str(dia+1)

        energia=0
        contaux=cont
        for hora in range(96):
            energia+=data[contaux]
            contaux+=1
        reg.energia=energia

        reg.reg00=data[cont+ 0]    
        reg.reg01=data[cont+ 1]    
        reg.reg02=data[cont+ 2]    
        reg.reg03=data[cont+ 3]    
        reg.reg04=data[cont+ 4]    
        reg.reg05=data[cont+ 5]    
        reg.reg06=data[cont+ 6]    
        reg.reg07=data[cont+ 7]    
        reg.reg08=data[cont+ 8]    
        reg.reg09=data[cont+ 9]    
        reg.reg10=data[cont+10]    
        reg.reg11=data[cont+11]    
        reg.reg12=data[cont+12]    
        reg.reg13=data[cont+13]    
        reg.reg14=data[cont+14]    
        reg.reg15=data[cont+15]    
        reg.reg16=data[cont+16]    
        reg.reg17=data[cont+17]    
        reg.reg18=data[cont+18]    
        reg.reg19=data[cont+19]    
        reg.reg20=data[cont+20]    
        reg.reg21=data[cont+21]    
        reg.reg22=data[cont+22]    
        reg.reg23=data[cont+23]    
        reg.reg24=data[cont+24]    
        reg.reg25=data[cont+25]    
        reg.reg26=data[cont+26]    
        reg.reg27=data[cont+27]    
        reg.reg28=data[cont+28]    
        reg.reg29=data[cont+29]    
        reg.reg30=data[cont+30]    
        reg.reg31=data[cont+31]    
        reg.reg32=data[cont+32]    
        reg.reg33=data[cont+33]    
        reg.reg34=data[cont+34]    
        reg.reg35=data[cont+35]    
        reg.reg36=data[cont+36]    
        reg.reg37=data[cont+37]    
        reg.reg38=data[cont+38]    
        reg.reg39=data[cont+39]    
        reg.reg40=data[cont+40]    
        reg.reg41=data[cont+41]    
        reg.reg42=data[cont+42]    
        reg.reg43=data[cont+43]    
        reg.reg44=data[cont+44]    
        reg.reg45=data[cont+45]    
        reg.reg46=data[cont+46]    
        reg.reg47=data[cont+47]    
        reg.reg48=data[cont+48]    
        reg.reg49=data[cont+49]    
        reg.reg50=data[cont+50]    
        reg.reg51=data[cont+51]    
        reg.reg52=data[cont+52]    
        reg.reg53=data[cont+53]    
        reg.reg54=data[cont+54]    
        reg.reg55=data[cont+55]    
        reg.reg56=data[cont+56]    
        reg.reg57=data[cont+57]    
        reg.reg58=data[cont+58]    
        reg.reg59=data[cont+59]    
        reg.reg60=data[cont+60]    
        reg.reg61=data[cont+61]    
        reg.reg62=data[cont+62]    
        reg.reg63=data[cont+63]    
        reg.reg64=data[cont+64]    
        reg.reg65=data[cont+65]    
        reg.reg66=data[cont+66]    
        reg.reg67=data[cont+67]    
        reg.reg68=data[cont+68]    
        reg.reg69=data[cont+69]    
        reg.reg70=data[cont+70]    
        reg.reg71=data[cont+71]    
        reg.reg72=data[cont+72]    
        reg.reg73=data[cont+73]    
        reg.reg74=data[cont+74]    
        reg.reg75=data[cont+75]    
        reg.reg76=data[cont+76]    
        reg.reg77=data[cont+77]    
        reg.reg78=data[cont+78]    
        reg.reg79=data[cont+79]    
        reg.reg80=data[cont+80]    
        reg.reg81=data[cont+81]    
        reg.reg82=data[cont+82]    
        reg.reg83=data[cont+83]    
        reg.reg84=data[cont+84]    
        reg.reg85=data[cont+85]    
        reg.reg86=data[cont+86]    
        reg.reg87=data[cont+87]    
        reg.reg88=data[cont+88]    
        reg.reg89=data[cont+89]    
        reg.reg90=data[cont+90]    
        reg.reg91=data[cont+91]    
        reg.reg92=data[cont+92]    
        reg.reg93=data[cont+93]    
        reg.reg94=data[cont+94]    
        reg.reg95=data[cont+95]    

        reg.usuario0=login
        reg.fecha0  =datetime.now()

        reg.save()    
        cont+=96

def BDobtieneregistroPot(miper,ubicacion,cods,tipo,potencia):    #Obtiene el registro de potencia en la ubicacion señalada
            # tipo=1: Registro  tipo=2: Con.Libre3eros  tipo=3: Reparto
        y=[]

        for i in range(len(cods)):
            vec=[] # en el [0] va el codigo
            vec.append(cods[i].id)
            vec.append(0)
            y.append(vec)
        
            dia=int(ubicacion/96)+1
            per=(ubicacion%96)

            j=Registro.objects.get(mesano=miper,dia=dia,codigo=cods[i].id,tipo=tipo,version=0,potencia=potencia) #jalar todos los datos

            vec=[]
            vec.append(  j.reg00)
            vec.append(  j.reg01)
            vec.append(  j.reg02)
            vec.append(  j.reg03)
            vec.append(  j.reg04)
            vec.append(  j.reg05)
            vec.append(  j.reg06)
            vec.append(  j.reg07)
            vec.append(  j.reg08)
            vec.append(  j.reg09)
            vec.append(  j.reg10)
            vec.append(  j.reg11)
            vec.append(  j.reg12)
            vec.append(  j.reg13)
            vec.append(  j.reg14)
            vec.append(  j.reg15)
            vec.append(  j.reg16)
            vec.append(  j.reg17)
            vec.append(  j.reg18)
            vec.append(  j.reg19)
            vec.append(  j.reg20)
            vec.append(  j.reg21)
            vec.append(  j.reg22)
            vec.append(  j.reg23)
            vec.append(  j.reg24)
            vec.append(  j.reg25)
            vec.append(  j.reg26)
            vec.append(  j.reg27)
            vec.append(  j.reg28)
            vec.append(  j.reg29)
            vec.append(  j.reg30)
            vec.append(  j.reg31)
            vec.append(  j.reg32)
            vec.append(  j.reg33)
            vec.append(  j.reg34)
            vec.append(  j.reg35)
            vec.append(  j.reg36)
            vec.append(  j.reg37)
            vec.append(  j.reg38)
            vec.append(  j.reg39)
            vec.append(  j.reg40)
            vec.append(  j.reg41)
            vec.append(  j.reg42)
            vec.append(  j.reg43)
            vec.append(  j.reg44)
            vec.append(  j.reg45)
            vec.append(  j.reg46)
            vec.append(  j.reg47)
            vec.append(  j.reg48)
            vec.append(  j.reg49)
            vec.append(  j.reg50)
            vec.append(  j.reg51)
            vec.append(  j.reg52)
            vec.append(  j.reg53)
            vec.append(  j.reg54)
            vec.append(  j.reg55)
            vec.append(  j.reg56)
            vec.append(  j.reg57)
            vec.append(  j.reg58)
            vec.append(  j.reg59)
            vec.append(  j.reg60)
            vec.append(  j.reg61)
            vec.append(  j.reg62)
            vec.append(  j.reg63)
            vec.append(  j.reg64)
            vec.append(  j.reg65)
            vec.append(  j.reg66)
            vec.append(  j.reg67)
            vec.append(  j.reg68)
            vec.append(  j.reg69)
            vec.append(  j.reg70)
            vec.append(  j.reg71)
            vec.append(  j.reg72)
            vec.append(  j.reg73)
            vec.append(  j.reg74)
            vec.append(  j.reg75)
            vec.append(  j.reg76)
            vec.append(  j.reg77)
            vec.append(  j.reg78)
            vec.append(  j.reg79)
            vec.append(  j.reg80)
            vec.append(  j.reg81)
            vec.append(  j.reg82)
            vec.append(  j.reg83)
            vec.append(  j.reg84)
            vec.append(  j.reg85)
            vec.append(  j.reg86)
            vec.append(  j.reg87)
            vec.append(  j.reg88)
            vec.append(  j.reg89)
            vec.append(  j.reg90)
            vec.append(  j.reg91)
            vec.append(  j.reg92)
            vec.append(  j.reg93)
            vec.append(  j.reg94)                                                
            vec.append(  j.reg95)

            y[i][1]=vec[per]
     
        return y


def BDobtieneregistro(miper,cods,tipo,potencia):    #siempre devuelve31 dias
            # tipo=1: Registro  tipo=2: Con.Libre3eros  tipo=3: Reparto
    y=[]

    #Repartos hace otro reordenamiento
        # en cods: esta listado barras
        # en potencia: se señala si es regulado o libre

    for i in range(len(cods)):

        vec=[] # en el [0] va el codigo
        vec.append(cods[i].id)

        for d in range(2976): vec.append(0)
        y.append(vec)
        
        if (tipo==1 or tipo==2): #Registros o Clientes Libres 3eros
            reg=Registro.objects.filter(mesano=miper,codigo=cods[i].id,tipo=tipo,version=0,potencia=potencia) #jalar todos los datos
        if (tipo==3):#Repartos
            reg=Registro.objects.filter(mesano=miper,codigo=potencia,barra=cods[i].id,tipo=tipo,version=0) #jalar todos los datos

        for j in reg:
            vec=[]
            vec.append(  j.reg00)
            vec.append(  j.reg01)
            vec.append(  j.reg02)
            vec.append(  j.reg03)
            vec.append(  j.reg04)
            vec.append(  j.reg05)
            vec.append(  j.reg06)
            vec.append(  j.reg07)
            vec.append(  j.reg08)
            vec.append(  j.reg09)
            vec.append(  j.reg10)
            vec.append(  j.reg11)
            vec.append(  j.reg12)
            vec.append(  j.reg13)
            vec.append(  j.reg14)
            vec.append(  j.reg15)
            vec.append(  j.reg16)
            vec.append(  j.reg17)
            vec.append(  j.reg18)
            vec.append(  j.reg19)
            vec.append(  j.reg20)
            vec.append(  j.reg21)
            vec.append(  j.reg22)
            vec.append(  j.reg23)
            vec.append(  j.reg24)
            vec.append(  j.reg25)
            vec.append(  j.reg26)
            vec.append(  j.reg27)
            vec.append(  j.reg28)
            vec.append(  j.reg29)
            vec.append(  j.reg30)
            vec.append(  j.reg31)
            vec.append(  j.reg32)
            vec.append(  j.reg33)
            vec.append(  j.reg34)
            vec.append(  j.reg35)
            vec.append(  j.reg36)
            vec.append(  j.reg37)
            vec.append(  j.reg38)
            vec.append(  j.reg39)
            vec.append(  j.reg40)
            vec.append(  j.reg41)
            vec.append(  j.reg42)
            vec.append(  j.reg43)
            vec.append(  j.reg44)
            vec.append(  j.reg45)
            vec.append(  j.reg46)
            vec.append(  j.reg47)
            vec.append(  j.reg48)
            vec.append(  j.reg49)
            vec.append(  j.reg50)
            vec.append(  j.reg51)
            vec.append(  j.reg52)
            vec.append(  j.reg53)
            vec.append(  j.reg54)
            vec.append(  j.reg55)
            vec.append(  j.reg56)
            vec.append(  j.reg57)
            vec.append(  j.reg58)
            vec.append(  j.reg59)
            vec.append(  j.reg60)
            vec.append(  j.reg61)
            vec.append(  j.reg62)
            vec.append(  j.reg63)
            vec.append(  j.reg64)
            vec.append(  j.reg65)
            vec.append(  j.reg66)
            vec.append(  j.reg67)
            vec.append(  j.reg68)
            vec.append(  j.reg69)
            vec.append(  j.reg70)
            vec.append(  j.reg71)
            vec.append(  j.reg72)
            vec.append(  j.reg73)
            vec.append(  j.reg74)
            vec.append(  j.reg75)
            vec.append(  j.reg76)
            vec.append(  j.reg77)
            vec.append(  j.reg78)
            vec.append(  j.reg79)
            vec.append(  j.reg80)
            vec.append(  j.reg81)
            vec.append(  j.reg82)
            vec.append(  j.reg83)
            vec.append(  j.reg84)
            vec.append(  j.reg85)
            vec.append(  j.reg86)
            vec.append(  j.reg87)
            vec.append(  j.reg88)
            vec.append(  j.reg89)
            vec.append(  j.reg90)
            vec.append(  j.reg91)
            vec.append(  j.reg92)
            vec.append(  j.reg93)
            vec.append(  j.reg94)                                                
            vec.append(  j.reg95)

            dia=j.dia
            pos=(dia-1)*96+1
            for z in range(96): y[i][pos+z]=vec[z]

    return y

def InspeccionarsihaydatosEnergia(miper,cod,tipo,barra,ver):
    # inspeccionar todos los datos existentes
    try:
        reg=Registro.objects.get(mesano=miper,codigo=cod.id,version=0,tipo=tipo,barra=barra,dia=1,potencia=False) #ver si tiene un dia
        datos=True
    except Registro.DoesNotExist:
        datos=False
    return datos
def InspeccionarsihaydatosPotencia(miper,cod,tipo,barra,ver):
    # inspeccionar todos los datos existentes
    try:
        reg=Registro.objects.get(mesano=miper,codigo=cod.id,version=0,tipo=tipo,barra=barra,dia=1,potencia=True) #ver si tiene un dia
        datos=True
    except Registro.DoesNotExist:
        datos=False
    return datos


def Excelobtenerregistro(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    nombres = wb.sheetnames


    data=[]    
    #********************************** INGRESAR A BASE DE DATOS ARCHIVO EXCEL *************
    if ('Registros' in nombres):
        s1 = wb["Registros"]
            

        fila=0

        for row in s1.iter_rows():

            if (fila==0):  #codigos

                codigoexcel=[]
                col=0
                ncol=0
                for cell in row:
                    if (col>0):
                        if (cell.value)!=None:
                            codigoexcel.append(str(cell.value))
                            registros=[]
                            registros.append(cell.value) #en la posicion 0 va el  codigo leido
                            data.append(registros)
                            ncol+=1
                    col+=1



            if (fila>0 and fila<2977): #valores
                for col in range(1,(ncol+1)):
                        try:
                            val=float(row[col].value)
                        except:    
                            val=0
                        data[col-1].append(val*0.001)
            fila+=1            

    return data

def ObtieneEnergia(barraBD,multiplicador,r):  #viene un registro con un reparto
    factor=[]
    pos=0
    # dia1
    for d in range(31):
        if (d==0): 
            fhp = r.porcentaje01 
            ffp = r.porcentajefp01
        if (d==1): 
            fhp = r.porcentaje02 
            ffp = r.porcentajefp02
        if (d==2): 
            fhp = r.porcentaje03 
            ffp = r.porcentajefp03
        if (d==3): 
            fhp = r.porcentaje04 
            ffp = r.porcentajefp04
        if (d==4): 
            fhp = r.porcentaje05 
            ffp = r.porcentajefp05
        if (d==5): 
            fhp = r.porcentaje06 
            ffp = r.porcentajefp06
        if (d==6): 
            fhp = r.porcentaje07 
            ffp = r.porcentajefp07
        if (d==7): 
            fhp = r.porcentaje08 
            ffp = r.porcentajefp08
        if (d==8): 
            fhp = r.porcentaje09 
            ffp = r.porcentajefp09
        if (d==9): 
            fhp = r.porcentaje10 
            ffp = r.porcentajefp10
        if (d==10): 
            fhp = r.porcentaje11 
            ffp = r.porcentajefp11
        if (d==11): 
            fhp = r.porcentaje12 
            ffp = r.porcentajefp12
        if (d==12): 
            fhp = r.porcentaje13 
            ffp = r.porcentajefp13
        if (d==13): 
            fhp = r.porcentaje14 
            ffp = r.porcentajefp14
        if (d==14): 
            fhp = r.porcentaje15 
            ffp = r.porcentajefp15
        if (d==15): 
            fhp = r.porcentaje16 
            ffp = r.porcentajefp16
        if (d==16): 
            fhp = r.porcentaje17 
            ffp = r.porcentajefp17
        if (d==17): 
            fhp = r.porcentaje18 
            ffp = r.porcentajefp18
        if (d==18): 
            fhp = r.porcentaje19 
            ffp = r.porcentajefp19
        if (d==19): 
            fhp = r.porcentaje20 
            ffp = r.porcentajefp20
        if (d==20): 
            fhp = r.porcentaje21 
            ffp = r.porcentajefp21
        if (d==21): 
            fhp = r.porcentaje22 
            ffp = r.porcentajefp22
        if (d==22): 
            fhp = r.porcentaje23 
            ffp = r.porcentajefp23
        if (d==23): 
            fhp = r.porcentaje24 
            ffp = r.porcentajefp24
        if (d==24): 
            fhp = r.porcentaje25 
            ffp = r.porcentajefp25
        if (d==25): 
            fhp = r.porcentaje26 
            ffp = r.porcentajefp26
        if (d==26): 
            fhp = r.porcentaje27 
            ffp = r.porcentajefp27
        if (d==27): 
            fhp = r.porcentaje28 
            ffp = r.porcentajefp28
        if (d==28): 
            fhp = r.porcentaje29 
            ffp = r.porcentajefp29
        if (d==29): 
            fhp = r.porcentaje30 
            ffp = r.porcentajefp30
        if (d==30): 
            fhp = r.porcentaje31 
            ffp = r.porcentajefp31

        for p in range(96):
            val=fhp if (72<=p and p<=91) else ffp
            factor.append(val)
    vector=[]
    vector.append(0)
    for i in range(2976):   vector.append(barraBD[i+1]*factor[i]*multiplicador)
    return vector